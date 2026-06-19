"""Fee analysis script.

Reads fee data from provided params, calculates effective fees after GST/RITC,
and writes an Excel file with Fee Summary and Benchmark Comparison sheets.

CLI usage:
    python fee_analysis.py --params '{"management_fee": 0.55, "performance_fee": 0}' \
                           --output /path/to/output.xlsx \
                           --project <project_id>
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path


def run(params: dict, output_path: Path, project_id: str) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    # ── Fee inputs ──────────────────────────────────────────────────────────
    management_fee: float = float(params.get("management_fee", 0.0))  # % p.a.
    performance_fee: float = float(params.get("performance_fee", 0.0))  # % p.a.
    indirect_costs: float = float(params.get("indirect_costs", 0.0))   # % p.a.
    buy_sell_spread: float = float(params.get("buy_sell_spread", 0.0))  # % per transaction
    investment_amount: float = float(params.get("investment_amount", 100_000.0))

    # GST = 10%; RITC (reduced input tax credit) = 55% of GST for financial services
    GST_RATE = 0.10
    RITC_RATE = 0.55  # typical RITC for managed funds
    net_gst_factor = 1 + GST_RATE * (1 - RITC_RATE)  # 1.045

    management_fee_inc_gst = management_fee * net_gst_factor
    performance_fee_inc_gst = performance_fee * net_gst_factor
    total_effective_fee = management_fee_inc_gst + performance_fee_inc_gst + indirect_costs

    # Dollar amounts
    mgmt_dollar = investment_amount * management_fee_inc_gst / 100
    perf_dollar = investment_amount * performance_fee_inc_gst / 100
    indirect_dollar = investment_amount * indirect_costs / 100
    total_dollar = investment_amount * total_effective_fee / 100

    # Benchmark (ASX 300 typical passive fee ~0.07%)
    benchmark_fee = 0.07
    fee_premium = total_effective_fee - benchmark_fee

    # ── Workbook ─────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ── Sheet 1: Fee Summary ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Fee Summary"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)

    def _header(ws, row: int, col: int, value: str) -> None:
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    def _bold(ws, row: int, col: int, value) -> None:
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = bold

    # Title
    ws1["A1"] = "Fee Analysis Report"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A2"] = f"Project ID: {project_id}"
    ws1["A3"] = f"Investment Amount: ${investment_amount:,.2f}"

    # Headers
    _header(ws1, 5, 1, "Fee Component")
    _header(ws1, 5, 2, "% p.a. (excl. GST)")
    _header(ws1, 5, 3, "Net GST Factor")
    _header(ws1, 5, 4, "% p.a. (incl. net GST)")
    _header(ws1, 5, 5, f"$ p.a. on ${investment_amount:,.0f}")

    # Rows
    rows = [
        ("Management Fee",      management_fee,  net_gst_factor, management_fee_inc_gst,  mgmt_dollar),
        ("Performance Fee",     performance_fee, net_gst_factor, performance_fee_inc_gst, perf_dollar),
        ("Indirect Costs (ICR)", indirect_costs,  1.0,            indirect_costs,          indirect_dollar),
    ]
    for i, (label, excl, factor, incl, dollar) in enumerate(rows, start=6):
        ws1.cell(row=i, column=1, value=label)
        ws1.cell(row=i, column=2, value=round(excl, 4))
        ws1.cell(row=i, column=3, value=round(factor, 4))
        ws1.cell(row=i, column=4, value=round(incl, 4))
        ws1.cell(row=i, column=5, value=round(dollar, 2))

    # Total
    total_row = len(rows) + 6
    _bold(ws1, total_row, 1, "Total Effective Fee")
    _bold(ws1, total_row, 4, round(total_effective_fee, 4))
    _bold(ws1, total_row, 5, round(total_dollar, 2))

    # Buy/sell spread note
    ws1.cell(row=total_row + 2, column=1, value="Buy/Sell Spread (per transaction):")
    ws1.cell(row=total_row + 2, column=2, value=f"{buy_sell_spread:.4f}%")

    # Column widths
    for col, width in zip("ABCDE", [30, 22, 18, 26, 24]):
        ws1.column_dimensions[get_column_letter(ord(col) - ord("A") + 1)].width = width

    # ── Sheet 2: Benchmark Comparison ────────────────────────────────────────
    ws2 = wb.create_sheet("Benchmark Comparison")

    _header(ws2, 1, 1, "Metric")
    _header(ws2, 1, 2, "This Fund")
    _header(ws2, 1, 3, "Passive Benchmark")
    _header(ws2, 1, 4, "Premium / (Discount)")

    comparison_rows = [
        ("Total Effective Fee (% p.a.)", round(total_effective_fee, 4), round(benchmark_fee, 4), round(fee_premium, 4)),
        ("Annual Cost on $100k",         round(total_effective_fee * 1000, 2), round(benchmark_fee * 1000, 2), round(fee_premium * 1000, 2)),
    ]
    for i, (label, fund_val, bench_val, premium) in enumerate(comparison_rows, start=2):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=fund_val)
        ws2.cell(row=i, column=3, value=bench_val)
        cell = ws2.cell(row=i, column=4, value=premium)
        if premium > 0:
            cell.font = Font(color="FF0000")  # red = more expensive
        else:
            cell.font = Font(color="00AA00")  # green = cheaper

    for col, width in zip("ABCD", [35, 18, 22, 24]):
        ws2.column_dimensions[get_column_letter(ord(col) - ord("A") + 1)].width = width

    # ── Save ─────────────────────────────────────────────────────────────────
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Fee analysis written to {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Fee analysis script")
    parser.add_argument("--params", required=True, help="JSON params string")
    parser.add_argument("--output", required=True, help="Output file path (.xlsx)")
    parser.add_argument("--project", required=True, help="Project ID")
    args = parser.parse_args()

    params = json.loads(args.params)
    run(params=params, output_path=Path(args.output), project_id=args.project)


if __name__ == "__main__":
    main()
