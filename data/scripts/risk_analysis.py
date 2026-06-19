"""Risk analysis script.

Calculates concentration risk flags and volatility proxies from holdings/returns data.
Writes a CSV summary and an Excel detail file.

CLI usage:
    python risk_analysis.py \
        --params '{"holdings": [...], "monthly_returns": [0.012, -0.008, ...]}' \
        --output /path/to/output.xlsx \
        --project <project_id>
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from collections import defaultdict
from pathlib import Path


def _std_dev(values: list[float]) -> float:
    """Population standard deviation."""
    if len(values) < 2:
        return 0.0
    mean = sum(values) / len(values)
    variance = sum((x - mean) ** 2 for x in values) / (len(values) - 1)
    return math.sqrt(variance)


def _max_drawdown(returns: list[float]) -> float:
    """Maximum peak-to-trough drawdown from a list of period returns."""
    if not returns:
        return 0.0
    peak = 1.0
    nav = 1.0
    max_dd = 0.0
    for r in returns:
        nav *= 1 + r
        if nav > peak:
            peak = nav
        dd = (peak - nav) / peak
        if dd > max_dd:
            max_dd = dd
    return max_dd


def run(params: dict, output_path: Path, project_id: str) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    holdings: list[dict] = params.get("holdings", [])
    monthly_returns: list[float] = [float(r) for r in params.get("monthly_returns", [])]

    # ── Concentration risk flags ──────────────────────────────────────────────
    total_weight = sum(float(h.get("weight", 0)) for h in holdings) or 1.0

    asset_class_weights: dict[str, float] = defaultdict(float)
    for h in holdings:
        asset_class_weights[h.get("asset_class", "Unknown")] += float(h.get("weight", 0))

    flags: list[dict] = []

    # Flag individual holdings > 10%
    for h in holdings:
        pct = float(h.get("weight", 0)) / total_weight * 100
        if pct > 10:
            flags.append({
                "flag_type": "CONCENTRATION",
                "severity": "HIGH" if pct > 20 else "MEDIUM",
                "description": f"{h.get('name', 'Unknown')} has {pct:.1f}% weight (threshold: 10%)",
            })

    # Flag single asset class > 70%
    for ac, weight in asset_class_weights.items():
        pct = weight / total_weight * 100
        if pct > 70:
            flags.append({
                "flag_type": "ASSET_CLASS_CONCENTRATION",
                "severity": "HIGH",
                "description": f"{ac} represents {pct:.1f}% of portfolio (threshold: 70%)",
            })

    # ── Volatility proxies ────────────────────────────────────────────────────
    annualised_vol = _std_dev(monthly_returns) * math.sqrt(12) * 100 if monthly_returns else None
    max_dd = _max_drawdown(monthly_returns) * 100 if monthly_returns else None
    avg_monthly_return = sum(monthly_returns) / len(monthly_returns) * 100 if monthly_returns else None

    # Flag high volatility
    if annualised_vol is not None and annualised_vol > 20:
        flags.append({
            "flag_type": "HIGH_VOLATILITY",
            "severity": "MEDIUM",
            "description": f"Annualised volatility of {annualised_vol:.1f}% exceeds 20% threshold",
        })

    if max_dd is not None and max_dd > 30:
        flags.append({
            "flag_type": "DRAWDOWN",
            "severity": "HIGH",
            "description": f"Maximum drawdown of {max_dd:.1f}% exceeds 30% threshold",
        })

    # ── Workbook ─────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    red_font = Font(color="FF0000", bold=True)
    amber_font = Font(color="FF8C00", bold=True)
    green_font = Font(color="00AA00")
    bold = Font(bold=True)

    def _header(ws, row: int, col: int, value: str) -> None:
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # ── Sheet 1: Risk Summary ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Risk Summary"

    ws1["A1"] = "Risk Analysis Report"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A2"] = f"Project ID: {project_id}"

    _header(ws1, 4, 1, "Metric")
    _header(ws1, 4, 2, "Value")
    _header(ws1, 4, 3, "Status")

    summary_rows: list[tuple] = [
        ("Number of Holdings", len(holdings), ""),
        ("Annualised Volatility (%)", f"{annualised_vol:.2f}" if annualised_vol is not None else "N/A",
         "HIGH" if (annualised_vol or 0) > 20 else "OK"),
        ("Max Drawdown (%)", f"{max_dd:.2f}" if max_dd is not None else "N/A",
         "HIGH" if (max_dd or 0) > 30 else "OK"),
        ("Avg Monthly Return (%)", f"{avg_monthly_return:.4f}" if avg_monthly_return is not None else "N/A", ""),
        ("Concentration Flags", len([f for f in flags if f["flag_type"] == "CONCENTRATION"]),
         "REVIEW" if any(f["flag_type"] == "CONCENTRATION" for f in flags) else "OK"),
    ]

    for i, (label, value, status) in enumerate(summary_rows, start=5):
        ws1.cell(row=i, column=1, value=label)
        ws1.cell(row=i, column=2, value=value)
        status_cell = ws1.cell(row=i, column=3, value=status)
        if status == "HIGH":
            status_cell.font = red_font
        elif status == "REVIEW":
            status_cell.font = amber_font
        elif status == "OK":
            status_cell.font = green_font

    for col, width in zip("ABC", [35, 20, 12]):
        ws1.column_dimensions[get_column_letter(ord(col) - ord("A") + 1)].width = width

    # ── Sheet 2: Risk Flags ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Risk Flags")

    if flags:
        _header(ws2, 1, 1, "Flag Type")
        _header(ws2, 1, 2, "Severity")
        _header(ws2, 1, 3, "Description")

        for i, flag in enumerate(flags, start=2):
            ws2.cell(row=i, column=1, value=flag["flag_type"])
            sev_cell = ws2.cell(row=i, column=2, value=flag["severity"])
            if flag["severity"] == "HIGH":
                sev_cell.font = red_font
            elif flag["severity"] == "MEDIUM":
                sev_cell.font = amber_font
            ws2.cell(row=i, column=3, value=flag["description"])
    else:
        ws2["A1"] = "No risk flags detected."
        ws2["A1"].font = green_font

    for col, width in zip("ABC", [25, 12, 70]):
        ws2.column_dimensions[get_column_letter(ord(col) - ord("A") + 1)].width = width

    # ── Sheet 3: Holdings Detail ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Holdings Detail")

    _header(ws3, 1, 1, "Name")
    _header(ws3, 1, 2, "Asset Class")
    _header(ws3, 1, 3, "Weight (%)")
    _header(ws3, 1, 4, "Portfolio Allocation (%)")

    for i, h in enumerate(sorted(holdings, key=lambda x: float(x.get("weight", 0)), reverse=True), start=2):
        pct = float(h.get("weight", 0)) / total_weight * 100
        ws3.cell(row=i, column=1, value=h.get("name", ""))
        ws3.cell(row=i, column=2, value=h.get("asset_class", ""))
        ws3.cell(row=i, column=3, value=round(float(h.get("weight", 0)), 4))
        ws3.cell(row=i, column=4, value=round(pct, 4))

    for col, width in zip("ABCD", [30, 20, 15, 22]):
        ws3.column_dimensions[get_column_letter(ord(col) - ord("A") + 1)].width = width

    # ── Save ─────────────────────────────────────────────────────────────────
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Risk analysis written to {output_path}")

    # Also write CSV summary
    csv_path = output_path.with_suffix(".csv")
    with csv_path.open("w", encoding="utf-8") as f:
        f.write("flag_type,severity,description\n")
        for flag in flags:
            desc = flag["description"].replace('"', '""')
            f.write(f'{flag["flag_type"]},{flag["severity"]},"{desc}"\n')
    print(f"CSV summary written to {csv_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Risk analysis script")
    parser.add_argument("--params", required=True, help="JSON params string")
    parser.add_argument("--output", required=True, help="Output file path (.xlsx)")
    parser.add_argument("--project", required=True, help="Project ID")
    args = parser.parse_args()

    params = json.loads(args.params)
    run(params=params, output_path=Path(args.output), project_id=args.project)


if __name__ == "__main__":
    main()
