"""Portfolio metrics script.

Calculates asset class allocation and concentration ratios from holdings data.
Writes an Excel file with Allocation Summary and Concentration sheets.

CLI usage:
    python portfolio_metrics.py \
        --params '{"holdings": [{"name": "BHP", "asset_class": "Equities", "weight": 12.5}]}' \
        --output /path/to/output.xlsx \
        --project <project_id>
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path


def run(params: dict, output_path: Path, project_id: str) -> None:
    try:
        import openpyxl
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    holdings: list[dict] = params.get("holdings", [])

    # ── Calculations ─────────────────────────────────────────────────────────
    # Asset class aggregation
    asset_class_weights: dict[str, float] = defaultdict(float)
    for h in holdings:
        ac = h.get("asset_class", "Unknown")
        w = float(h.get("weight", 0.0))
        asset_class_weights[ac] += w

    total_weight = sum(asset_class_weights.values()) or 1.0
    asset_class_pct = {ac: w / total_weight * 100 for ac, w in asset_class_weights.items()}

    # Top-N concentration
    sorted_holdings = sorted(holdings, key=lambda h: float(h.get("weight", 0)), reverse=True)
    top5_weight = sum(float(h.get("weight", 0)) for h in sorted_holdings[:5]) / total_weight * 100
    top10_weight = sum(float(h.get("weight", 0)) for h in sorted_holdings[:10]) / total_weight * 100
    largest_holding = sorted_holdings[0] if sorted_holdings else {}

    # ── Workbook ─────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)

    def _header(ws, row: int, col: int, value: str) -> None:
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # ── Sheet 1: Allocation Summary ───────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Allocation Summary"

    ws1["A1"] = "Portfolio Allocation Report"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A2"] = f"Project ID: {project_id}"
    ws1["A3"] = f"Total holdings: {len(holdings)}"

    _header(ws1, 5, 1, "Asset Class")
    _header(ws1, 5, 2, "Total Weight (%)")
    _header(ws1, 5, 3, "Allocation (%)")

    for i, (ac, weight) in enumerate(sorted(asset_class_pct.items(), key=lambda x: -x[1]), start=6):
        ws1.cell(row=i, column=1, value=ac)
        ws1.cell(row=i, column=2, value=round(asset_class_weights[ac], 2))
        ws1.cell(row=i, column=3, value=round(weight, 2))

    end_row = 5 + len(asset_class_pct)

    # Pie chart of allocations
    if len(asset_class_pct) > 0:
        chart = PieChart()
        chart.title = "Asset Class Allocation"
        chart.style = 10
        labels = Reference(ws1, min_col=1, min_row=6, max_row=end_row)
        data = Reference(ws1, min_col=3, min_row=5, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        ws1.add_chart(chart, "E5")

    for col, width in zip("ABC", [25, 20, 18]):
        ws1.column_dimensions[get_column_letter(ord(col) - ord("A") + 1)].width = width

    # ── Sheet 2: Concentration ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Concentration")

    ws2["A1"] = "Concentration Analysis"
    ws2["A1"].font = Font(bold=True, size=14)

    _header(ws2, 3, 1, "Metric")
    _header(ws2, 3, 2, "Value")

    concentration_rows = [
        ("Number of Holdings", len(holdings)),
        ("Top 5 Holdings Weight (%)", round(top5_weight, 2)),
        ("Top 10 Holdings Weight (%)", round(top10_weight, 2)),
        ("Largest Single Holding", largest_holding.get("name", "N/A")),
        ("Largest Holding Weight (%)", round(float(largest_holding.get("weight", 0)) / total_weight * 100, 2) if largest_holding else 0),
    ]
    for i, (label, value) in enumerate(concentration_rows, start=4):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=value)

    # Individual holdings table
    _header(ws2, 10, 1, "Rank")
    _header(ws2, 10, 2, "Name")
    _header(ws2, 10, 3, "Asset Class")
    _header(ws2, 10, 4, "Weight (%)")
    _header(ws2, 10, 5, "Portfolio Allocation (%)")

    for rank, h in enumerate(sorted_holdings, start=1):
        row = 10 + rank
        pct = float(h.get("weight", 0)) / total_weight * 100
        ws2.cell(row=row, column=1, value=rank)
        ws2.cell(row=row, column=2, value=h.get("name", ""))
        ws2.cell(row=row, column=3, value=h.get("asset_class", ""))
        ws2.cell(row=row, column=4, value=round(float(h.get("weight", 0)), 4))
        ws2.cell(row=row, column=5, value=round(pct, 4))

    for col, width in zip("ABCDE", [8, 30, 20, 15, 22]):
        ws2.column_dimensions[get_column_letter(ord(col) - ord("A") + 1)].width = width

    # ── Save ─────────────────────────────────────────────────────────────────
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Portfolio metrics written to {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Portfolio metrics script")
    parser.add_argument("--params", required=True, help="JSON params string")
    parser.add_argument("--output", required=True, help="Output file path (.xlsx)")
    parser.add_argument("--project", required=True, help="Project ID")
    args = parser.parse_args()

    params = json.loads(args.params)
    run(params=params, output_path=Path(args.output), project_id=args.project)


if __name__ == "__main__":
    main()
