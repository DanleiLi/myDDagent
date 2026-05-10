"""
Generate fee analysis for managed portfolios.
Output: Fee Analysis - [Portfolio Series] - [Date].xlsx with 3 sheets:
  1. Fee Summary — fee components with Excel formulas
  2. Component Table — per-fund fee breakdown
  3. Portfolio Holdings — portfolio holdings and allocation
"""

import sys
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

OUTPUT_DIR = r'C:\Users\Sara\Downloads\AIagentproject\.claude\output'
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─────────────────────────────────────────────
# Assumptions and Fee Data
# ─────────────────────────────────────────────
# NOTE:
#   - All fee values in TAX_DATA, PORTFOLIO_DATA and FUND_FEES are specified in percent (e.g., 0.10 = 10%).
#   - Be careful not to confuse basis points (1 bps = 0.01%) with percent values.
#   - im_fee_bps and re_fee_bps are in percent.

TAX_DATA = {
    'GST_RATE': 0.10,  # 10% GST
    'IM_RITC': 0.75,   # 75% RITC for Investment Management fees
    'RE_RITC': 0.55,   # 55% RITC for Rebate fees
}

PORTFOLIOS = [
    {
        'portfolio_name': 'BlackRock Balanced AU Managed Portfolio',
        'series': 'BlackRock',
        'im_fee_bps': 0.45,
        're_fee_bps': 0.08,
        'holdings': [
            {'apir': 'AMP9555AU', 'portfolio_id': 'NTH0001', 'portfolio_name': 'BlackRock Balanced AU Managed Portfolio', 'fund_name': 'AMP Australian Equity Index Fund', 'allocation': 0.04},
            {'apir': 'GHLD', 'portfolio_id': 'NTH0001', 'portfolio_name': 'BlackRock Balanced AU Managed Portfolio', 'fund_name': 'Global X Gold Bullion (Currency Hedged) ETF', 'allocation': 0.03},
            {'apir': 'PER1058AU', 'portfolio_id': 'NTH0001', 'portfolio_name': 'BlackRock Balanced AU Managed Portfolio', 'fund_name': 'Perpetual Diversified Income Fund - Class S', 'allocation': 0.14},
        ]
    },
]

FUND_FEES = {
    'AMP9555AU': {
        'mgmt': 0.15, 'cash_inv': 0.0, 'perf': 0.0, 'transaction': 0.15,
        'buy_spread': 0.10, 'sell_spread': 0.10, 'rebate': 0.05, 'pds_url': 'DEMO_PLACEHOLDER'
    }
}


# ─────────────────────────────────────────────
# Brand colours (hex)
# ─────────────────────────────────────────────


COLOURS = {
    'primary_blue': '0B1EEA',
    'violet': '3A0CA3',
    'dark_purple': '240046',
    'lavender': 'D38BFF',
    'soft_lavender': 'C77DFF',
    'amber': 'FFC000',
    'white': 'FFFFFF',
    'light_grey': 'F0F0F0',
}

# ─────────────────────────────────────────────
#  Functions
# ─────────────────────────────────────────────

def apply_header_style(cell, bg_colour=None, font_colour='FFFFFF'):
    """Apply header styling to a cell."""
    if bg_colour is None:
        bg_colour = COLOURS['primary_blue']
    cell.fill = PatternFill(start_color=bg_colour, end_color=bg_colour, fill_type='solid')
    cell.font = Font(bold=True, color=font_colour)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def apply_data_style(cell, bg_colour=None, number_format=None):
    """Apply data cell styling."""
    cell.alignment = Alignment(horizontal='left', vertical='center')
    if bg_colour:
        cell.fill = PatternFill(start_color=bg_colour, end_color=bg_colour, fill_type='solid')
    if number_format:
        cell.number_format = number_format

def apply_formula_style(cell, number_format='0.00%'):
    """Apply formula cell styling."""
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.number_format = number_format
    cell.fill = PatternFill(start_color=COLOURS['light_grey'], end_color=COLOURS['light_grey'], fill_type='solid')

def calculate_im_fee_bps(im_fee_bps, re_fee_bps):
    """Calculate IM+RE fee including GST and RITC (for console output sanity check)."""
    im_component = (im_fee_bps) * (1 + TAX_DATA['GST_RATE'] - TAX_DATA['GST_RATE'] * TAX_DATA['IM_RITC']) 
    re_component = (re_fee_bps) * (1 + TAX_DATA['GST_RATE'] - TAX_DATA['GST_RATE'] * TAX_DATA['RE_RITC']) 
    return (im_component + re_component) 

# ─────────────────────────────────────────────
# Workbook Creation
# ─────────────────────────────────────────────

wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Build portfolio ranges for formula references
portfolio_ranges = []   # list of (first_data_row, last_data_row) per portfolio
current_row = 2         # row 1 is the header in Portfolio Holdings / Component
for p in PORTFOLIOS:
    n = len(p['holdings'])
    portfolio_ranges.append((current_row, current_row + n - 1))
    current_row += n

# ═════════════════════════════════════════════
# SHEET 1: Portfolio Holdings
# ═════════════════════════════════════════════

ws_component_table = wb.create_sheet('Component', 1)

# Header row
headers_table = [
    'Unit ID', 'Unit Name', 'Management fees and costs %', 'Cash investment fee %',
    'Performance fees %', 'Gross transaction costs %', 'Buy spread %', 'Sell spread %',
    'Rebate %', 'PDS URL'
]
for col, header in enumerate(headers_table, 1):
    cell = ws_component_table.cell(row=1, column=col, value=header)
    apply_header_style(cell)

# Data rows (in same order as Portfolio Holdings sheet)
row = 2
for p in PORTFOLIOS:
    for holding in p['holdings']:
        apir = holding['apir']
        fee_data = FUND_FEES[apir]

        ws_component_table.cell(row=row, column=1, value=apir)
        ws_component_table.cell(row=row, column=2, value=holding['fund_name'])

        # Fee columns (3-9)
        fee_values = [
            fee_data['mgmt']/100,
            fee_data['cash_inv']/100,
            fee_data['perf']/100,
            fee_data['transaction']/100,
            fee_data['buy_spread']/100,
            fee_data['sell_spread']/100,
            fee_data['rebate']/100,
        ]

        for col, fee_value in enumerate(fee_values, 3):
            cell = ws_component_table.cell(row=row, column=col, value=fee_value)
            apply_data_style(cell, number_format='0.00%')

        # PDS URL
        cell_url = ws_component_table.cell(row=row, column=10, value=fee_data['pds_url'])
        apply_data_style(cell_url)
        row += 1
   

# Column widths
ws_component_table.column_dimensions['A'].width = 12
ws_component_table.column_dimensions['B'].width = 20
for col in range(3, 10):
    ws_component_table.column_dimensions[get_column_letter(col)].width = 14
ws_component_table.column_dimensions['J'].width = 18

# ═════════════════════════════════════════════
# SHEET 3: Portfolio Holdings
# ═════════════════════════════════════════════

ws_component = wb.create_sheet('Portfolio Holdings', 2)

# Header row
headers = ['Portfolio ID', 'Portfolio Name', 'Unit ID', 'Allocation %']
for col, header in enumerate(headers, 1):
    cell = ws_component.cell(row=1, column=col, value=header)
    apply_header_style(cell)

# Data rows
row = 2
for p in PORTFOLIOS:
    for holding in p['holdings']:
        ws_component.cell(row=row, column=1, value=holding['portfolio_id'])
        ws_component.cell(row=row, column=2, value=holding['portfolio_name'])
        ws_component.cell(row=row, column=3, value=holding['apir'])
        cell_alloc = ws_component.cell(row=row, column=4, value=holding['allocation'])
        apply_data_style(cell_alloc, number_format='0.00%')
        row += 1

# Column widths
ws_component.column_dimensions['A'].width = 15
ws_component.column_dimensions['B'].width = 25
ws_component.column_dimensions['C'].width = 12
ws_component.column_dimensions['D'].width = 12

# ═════════════════════════════════════════════
# SHEET 1: Fee Summary (with Formulas)
# ═════════════════════════════════════════════

ws_fee_summary = wb.create_sheet('Fee Summary', 0)

# Parameters labels in column A
ws_fee_summary.cell(row=1, column=1, value='IM Fee % p.a.')
ws_fee_summary.cell(row=2, column=1, value='RE Fee % p.a.')

# Fee Summary Matrix Header
cell_a3 = ws_fee_summary.cell(row=3, column=1, value='Fee Component')
apply_header_style(cell_a3)

# Tax data for formula construction
gst = TAX_DATA['GST_RATE']
im_ritc = TAX_DATA['IM_RITC']
re_ritc = TAX_DATA['RE_RITC']

# Process each portfolio
for i, p in enumerate(PORTFOLIOS):
    col_idx = i + 2   # B=2, C=3, D=4, ...
    col = get_column_letter(col_idx)

    # Parameter rows (rows 1-2) with IM and RE fees
    im_cell = ws_fee_summary.cell(row=1, column=col_idx, value=p['im_fee_bps'] / 100)
    im_cell.number_format = '0.00%'

    re_cell = ws_fee_summary.cell(row=2, column=col_idx, value=p['re_fee_bps'] / 100)
    re_cell.number_format = '0.00%'

    # Header row (row 3) with portfolio name
    cell_hdr = ws_fee_summary.cell(row=3, column=col_idx, value=p['portfolio_name'])
    apply_header_style(cell_hdr)

    # Row ranges for this portfolio in Portfolio Holdings / Component
    first, last = portfolio_ranges[i]

    # Fee component formulas (rows 4-12)
    fee_components = [
        ('Investment management fees % p.a.',
         f'=({col}$1*(1+{gst}-{gst}*{im_ritc})+{col}$2*(1+{gst}-{gst}*{re_ritc}))'),
        ('Rebate % p.a.',
         f"=SUMPRODUCT('Portfolio Holdings'!$D${first}:$D${last},'Component'!$I${first}:$I${last})"),
        ('Estimated underlying management fees % p.a.',
         f"=SUMPRODUCT('Portfolio Holdings'!$D${first}:$D${last},'Component'!$C${first}:$C${last})-{col}5"),
        ('Estimated managed portfolio cash investment fee % p.a.',
         f"=SUMPRODUCT('Portfolio Holdings'!$D${first}:$D${last},'Component'!$D${first}:$D${last})"),
        ('Portfolio performance fee % p.a.', '=0'),
        ('Estimated underlying performance fee % p.a.',
         f"=SUMPRODUCT('Portfolio Holdings'!$D${first}:$D${last},'Component'!$E${first}:$E${last})"),
        ('Estimated gross transaction costs % p.a.',
         f"=SUMPRODUCT('Portfolio Holdings'!$D${first}:$D${last},'Component'!$F${first}:$F${last})"),
        ('Estimated underlying buy spread % p.a.',
         f"=SUMPRODUCT('Portfolio Holdings'!$D${first}:$D${last},'Component'!$G${first}:$G${last})"),
        ('Estimated underlying sell spread % p.a.',
         f"=SUMPRODUCT('Portfolio Holdings'!$D${first}:$D${last},'Component'!$H${first}:$H${last})"),
    ]

    for row_offset, (label, formula) in enumerate(fee_components, 4):
        # Write labels only once (in column A)
        if i == 0:
            cell_label = ws_fee_summary.cell(row=row_offset, column=1, value=label)
            cell_label.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Write formula for this portfolio
        cell_formula = ws_fee_summary.cell(row=row_offset, column=col_idx)
        cell_formula.value = formula
        apply_formula_style(cell_formula)

        # Highlight key rows: IM fee (row 4) and net underlying mgmt fee (row 6)
        if row_offset in [4, 6]:
            cell_formula.fill = PatternFill(start_color=COLOURS['white'], end_color=COLOURS['white'], fill_type='solid')
            if i == 0:
                cell_label.fill = PatternFill(start_color=COLOURS['white'], end_color=COLOURS['white'], fill_type='solid')
                cell_label.font = Font(color='000000')

# Column widths
ws_fee_summary.column_dimensions['A'].width = 50
for i in range(len(PORTFOLIOS)):
    ws_fee_summary.column_dimensions[get_column_letter(i + 2)].width = 18

# ─────────────────────────────────────────────
# Save Workbook
# ─────────────────────────────────────────────

all_series = PORTFOLIOS[0]['series'] if PORTFOLIOS else 'UnknownSeries'

output_filename = f"Fee Analysis - {all_series} - {datetime.now().strftime('%Y%m%d')}.xlsx"
output_path = os.path.join(OUTPUT_DIR, output_filename)

wb.save(output_path)

# Console output
print(f"✓ Fee analysis workbook generated successfully")
print(f"  Location: {output_path}")

