'''
This script reads an XLSX or CSV and saves each visible worksheet as JSON in wiki/. Ignoring formatted but empty rows and columns. Each JSON file contains the source file, sheet name, and rows of data. The script also prints a summary of sheets and row counts.
'''

import os
import csv
import json
from datetime import datetime
from openpyxl import load_workbook

##Parameters
file_path = "raw_document/IMQuestionnaire.xlsx"
output_dir = "wiki"
doc_label = 'IM Questionnaire'

## Code
os.makedirs(output_dir, exist_ok=True)

cleaned_data = {
    'source_file': file_path,
    'ingestion_timestamp': datetime.now().isoformat(),
    'cleanup_timestamp': datetime.now().isoformat(),
    'document_label': doc_label,
    'sheets': {}
}


def sanitize_sheet_name(name):
    safe_name = ''.join(c if c.isalnum() or c in (' ', '_', '-') else '_' for c in name).strip()
    safe_name = safe_name.replace(' ', '_')
    return safe_name or 'sheet'


def trim_empty_rows_and_columns(rows):
    if not rows:
        return []

    normalized = [[cell if cell is not None else '' for cell in row] for row in rows]
    row_has_value = [any(cell != '' for cell in row) for row in normalized]
    if not any(row_has_value):
        return []

    first_row = next(i for i, has in enumerate(row_has_value) if has)
    last_row = len(row_has_value) - 1 - next(i for i, has in enumerate(reversed(row_has_value)) if has)

    max_col = 0
    for row in normalized[first_row:last_row + 1]:
        for idx, cell in enumerate(row):
            if cell != '':
                max_col = max(max_col, idx + 1)

    trimmed = []
    for row in normalized[first_row:last_row + 1]:
        row_copy = list(row[:max_col])
        if len(row_copy) < max_col:
            row_copy.extend([''] * (max_col - len(row_copy)))
        trimmed.append(row_copy)

    return trimmed


def read_csv_file(path):
    rows = []
    with open(path, 'r', newline='', encoding='utf-8-sig') as csv_file:
        reader = csv.reader(csv_file)
        for row in reader:
            rows.append([cell if cell is not None else '' for cell in row])
    return trim_empty_rows_and_columns(rows)


if file_path.lower().endswith('.csv'):
    sheet_name = os.path.splitext(os.path.basename(file_path))[0]
    cleaned_data['sheets'][sheet_name] = read_csv_file(file_path)
else:
    wb = load_workbook(file_path, data_only=True)
    for sheet in wb.worksheets:
        if sheet.sheet_state != 'visible':
            continue

        sheet_name = sheet.title
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append([cell if cell is not None else '' for cell in row])
        cleaned_data['sheets'][sheet_name] = trim_empty_rows_and_columns(rows)

for sheet_name, rows in cleaned_data['sheets'].items():
    safe_name = sanitize_sheet_name(sheet_name)
    json_filename = f"{safe_name}.json"
    sheet_json_path = os.path.join(output_dir, json_filename)

    sheet_data = {
        'source_file': file_path,
        'sheet_name': sheet_name,
        'rows': rows            
    }
    with open(sheet_json_path, 'w', encoding='utf-8') as json_file:
        json.dump(sheet_data, json_file, indent=2, default=str)
    print(f"Saved sheet JSON: {sheet_json_path}")

for sheet_name, content in cleaned_data['sheets'].items():
    print(f"  {sheet_name}: {len(content)} rows")