"""
Generate fee analysis for managed portfolios.
Output: Fee Analysis - [Portfolio Series] - [Date].xlsx with 3 sheets:
  1. Fee Summary — fee components with Excel formulas
  2. Component Table — per-fund fee breakdown
  3. Portfolio Holdings — portfolio holdings and allocation

Usage:
  python generate_fee_analysis.py <config_json_path> [output_xlsx_dir]

  config_json_path: Path to JSON file with portfolios, fund_fees, tax_data, series_name
  output_xlsx_dir: Directory for Excel output (default: .claude/output relative to project root)
"""

import sys
import os
import json
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

# ─────────────────────────────────────────────
# Load Configuration from JSON
# ─────────────────────────────────────────────

def load_config(config_path):
    """Load portfolio and fee configuration from JSON file."""
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"Config file not found: {config_path}")
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)

    portfolios = config.get('portfolios', [])
    fund_fees = config.get('fund_fees', {})
    tax_data = config.get('tax_data') or {
        'GST_RATE': 0.10,
        'IM_RITC': 0.75,
        'RE_RITC': 0.55,
    }
    series_name = config.get('series_name', 'Portfolio Series')

    return portfolios, fund_fees, tax_data, series_name

# Get config path from command-line argument
if len(sys.argv) < 2:
    print("Error: Usage: python generate_fee_analysis.py <config_json_path> [output_xlsx_dir]")
    sys.exit(1)

CONFIG_PATH = sys.argv[1]
PORTFOLIOS, FUND_FEES, TAX_DATA, SERIES_NAME = load_config(CONFIG_PATH)

# Determine output directory
if len(sys.argv) > 2:
    OUTPUT_DIR = sys.argv[2]
else:
    # Default to .claude/output relative to project root (parent of .claude dir)
    script_dir = Path(__file__).parent
    project_root = script_dir.parent.parent.parent.parent  # go up 4 levels to project root
    OUTPUT_DIR = str(project_root / '.claude' / 'output')

os.makedirs(OUTPUT_DIR, exist_ok=True)


# ─────────────────────────────────────────────
# Brand colours (hex)
# ─────────────────────────────────────────────


COLOURS = {
    'primary_blue': '0B1EEA',
    'violet': '3A0CA3',
    'dark_purple': '240046',
    'lavender': 'D38BFF',
    'soft_lavender': 'C77DFF',
    'amber': 'FFC000',
    'white': 'FFFFFF',
    'light_grey': 'F0F0F0',
}

# ─────────────────────────────────────────────
#  Functions
# ─────────────────────────────────────────────

def apply_header_style(cell, bg_colour=None, font_colour='FFFFFF'):
    """Apply header styling to a cell."""
    if bg_colour is None:
        bg_colour = COLOURS['primary_blue']
    cell.fill = PatternFill(start_color=bg_colour, end_color=bg_colour, fill_type='solid')
    cell.font = Font(bold=True, color=font_colour)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def apply_data_style(cell, bg_colour=None, number_format=None):
    """Apply data cell styling."""
    cell.alignment = Alignment(horizontal='left', vertical='center')
    if bg_colour:
        cell.fill = PatternFill(start_color=bg_colour, end_color=bg_colour, fill_type='solid')
    if number_format:
        cell.number_format = number_format

def apply_formula_style(cell, number_format='0.00%'):
    """Apply formula cell styling."""
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.number_format = number_format
    cell.fill = PatternFill(start_color=COLOURS['light_grey'], end_color=COLOURS['light_grey'], fill_type='solid')

def calculate_im_fee_bps(im_fee_bps, re_fee_bps):
    """Calculate IM+RE fee including GST and RITC (for console output sanity check)."""
    im_component = (im_fee_bps) * (1 + TAX_DATA['GST_RATE'] - TAX_DATA['GST_RATE'] * TAX_DATA['IM_RITC']) 
    re_component = (re_fee_bps) * (1 + TAX_DATA['GST_RATE'] - TAX_DATA['GST_RATE'] * TAX_DATA['RE_RITC']) 
    return (im_component + re_component) 

# ─────────────────────────────────────────────
# Workbook Creation
# ─────────────────────────────────────────────

wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Build portfolio ranges for formula references
portfolio_ranges = []   # list of (first_data_row, last_data_row) per portfolio
current_row = 2         # row 1 is the header in Portfolio Holdings / Component
for p in PORTFOLIOS:
    n = len(p['holdings'])
    portfolio_ranges.append((current_row, current_row + n - 1))
    current_row += n

# ═════════════════════════════════════════════
# SHEET 1: Portfolio Holdings
# ═════════════════════════════════════════════

ws_component_table = wb.create_sheet('Component', 1)

# Header row
headers_table = [
    'Unit ID', 'Unit Name', 'Management fees and costs %', 'Cash investment fee %',
    'Performance fees %', 'Gross transaction costs %', 'Buy spread %', 'Sell spread %',
    'Rebate %', 'PDS URL'
]
for col, header in enumerate(headers_table, 1):
    cell = ws_component_table.cell(row=1, column=col, value=header)
    apply_header_style(cell)

# Data rows (in same order as Portfolio Holdings sheet)
row = 2
for p in PORTFOLIOS:
    for holding in p['holdings']:
        apir = holding['apir']
        fee_data = FUND_FEES[apir]

        ws_component_table.cell(row=row, column=1, value=apir)
        ws_component_table.cell(row=row, column=2, value=holding['fund_name'])

        # Fee columns (3-9)
        fee_values = [
            fee_data['mgmt']/100,
            fee_data['cash_inv']/100,
            fee_data['perf']/100,
            fee_data['transaction']/100,
            fee_data['buy_spread']/100,
            fee_data['sell_spread']/100,
            fee_data['rebate']/100,
        ]

        for col, fee_value in enumerate(fee_values, 3):
            cell = ws_component_table.cell(row=row, column=col, value=fee_value)
            apply_data_style(cell, number_format='0.00%')

        # PDS URL
        cell_url = ws_component_table.cell(row=row, column=10, value=fee_data['pds_url'])
        apply_data_style(cell_url)
        row += 1
   

# Column widths
ws_component_table.column_dimensions['A'].width = 12
ws_component_table.column_dimensions['B'].width = 20
for col in range(3, 10):
    ws_component_table.column_dimensions[get_column_letter(col)].width = 14
ws_component_table.column_dimensions['J'].width = 18

# ═════════════════════════════════════════════
# SHEET 3: Portfolio Holdings
# ═════════════════════════════════════════════

ws_component = wb.create_sheet('Portfolio Holdings', 2)

# Header row
headers = ['Portfolio ID', 'Portfolio Name', 'Unit ID', 'Allocation %']
for col, header in enumerate(headers, 1):
    cell = ws_component.cell(row=1, column=col, value=header)
    apply_header_style(cell)

# Data rows
row = 2
for p in PORTFOLIOS:
    for holding in p['holdings']:
        ws_component.cell(row=row, column=1, value=holding['portfolio_id'])
        ws_component.cell(row=row, column=2, value=holding['portfolio_name'])
        ws_component.cell(row=row, column=3, value=holding['apir'])
        cell_alloc = ws_component.cell(row=row, column=4, value=holding['allocation'])
        apply_data_style(cell_alloc, number_format='0.00%')
        row += 1

# Column widths
ws_component.column_dimensions['A'].width = 15
ws_component.column_dimensions['B'].width = 25
ws_component.column_dimensions['C'].width = 12
ws_component.column_dimensions['D'].width = 12

# ═════════════════════════════════════════════
# SHEET 1: Fee Summary (with Formulas)
# ═════════════════════════════════════════════

ws_fee_summary = wb.create_sheet('Fee Summary', 0)

# Parameters labels in column A
ws_fee_summary.cell(row=1, column=1, value='IM Fee % p.a.')
ws_fee_summary.cell(row=2, column=1, value='RE Fee % p.a.')

# Fee Summary Matrix Header
cell_a3 = ws_fee_summary.cell(row=3, column=1, value='Fee Component')
apply_header_style(cell_a3)

# Tax data for formula construction
gst = TAX_DATA['GST_RATE']
im_ritc = TAX_DATA['IM_RITC']
re_ritc = TAX_DATA['RE_RITC']

# Process each portfolio
for i, p in enumerate(PORTFOLIOS):
    col_idx = i + 2   # B=2, C=3, D=4, ...
    col = get_column_letter(col_idx)

    # Parameter rows (rows 1-2) with IM and RE fees
    im_cell = ws_fee_summary.cell(row=1, column=col_idx, value=p['im_fee_bps'] / 100)
    im_cell.number_format = '0.00%'

    re_cell = ws_fee_summary.cell(row=2, column=col_idx, value=p['re_fee_bps'] / 100)
    re_cell.number_format = '0.00%'

    # Header row (row 3) with portfolio name
    cell_hdr = ws_fee_summary.cell(row=3, column=col_idx, value=p['portfolio_name'])
    apply_header_style(cell_hdr)

    # Row ranges for this portfolio in Portfolio Holdings / Component
    first, last = portfolio_ranges[i]

    # Fee component formulas (rows 4-12)
    fee_components = [
        ('Investment management fees % p.a.',
         f'=({col}$1*(1+{gst}-{gst}*{im_ritc})+{col}$2*(1+{gst}-{gst}*{re_ritc}))'),
        ('Rebate % p.a.',
         f"=SUMPRODUCT('Portfolio Holdings'!$D${first}:$D${last},'Component'!$I${first}:$I${last})"),
        ('Estimated underlying management fees % p.a.',
         f"=SUMPRODUCT('Portfolio Holdings'!$D${first}:$D${last},'Component'!$C${first}:$C${last})-{col}5"),
        ('Estimated managed portfolio cash investment fee % p.a.',
         f"=SUMPRODUCT('Portfolio Holdings'!$D${first}:$D${last},'Component'!$D${first}:$D${last})"),
        ('Portfolio performance fee % p.a.', '=0'),
        ('Estimated underlying performance fee % p.a.',
         f"=SUMPRODUCT('Portfolio Holdings'!$D${first}:$D${last},'Component'!$E${first}:$E${last})"),
        ('Estimated gross transaction costs % p.a.',
         f"=SUMPRODUCT('Portfolio Holdings'!$D${first}:$D${last},'Component'!$F${first}:$F${last})"),
        ('Estimated underlying buy spread % p.a.',
         f"=SUMPRODUCT('Portfolio Holdings'!$D${first}:$D${last},'Component'!$G${first}:$G${last})"),
        ('Estimated underlying sell spread % p.a.',
         f"=SUMPRODUCT('Portfolio Holdings'!$D${first}:$D${last},'Component'!$H${first}:$H${last})"),
    ]

    for row_offset, (label, formula) in enumerate(fee_components, 4):
        # Write labels only once (in column A)
        if i == 0:
            cell_label = ws_fee_summary.cell(row=row_offset, column=1, value=label)
            cell_label.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Write formula for this portfolio
        cell_formula = ws_fee_summary.cell(row=row_offset, column=col_idx)
        cell_formula.value = formula
        apply_formula_style(cell_formula)

        # Highlight key rows: IM fee (row 4) and net underlying mgmt fee (row 6)
        if row_offset in [4, 6]:
            cell_formula.fill = PatternFill(start_color=COLOURS['white'], end_color=COLOURS['white'], fill_type='solid')
            if i == 0:
                cell_label.fill = PatternFill(start_color=COLOURS['white'], end_color=COLOURS['white'], fill_type='solid')
                cell_label.font = Font(color='000000')

# Column widths
ws_fee_summary.column_dimensions['A'].width = 50
for i in range(len(PORTFOLIOS)):
    ws_fee_summary.column_dimensions[get_column_letter(i + 2)].width = 18

# ─────────────────────────────────────────────
# Save Workbook
# ─────────────────────────────────────────────

output_filename = f"Fee Analysis - {SERIES_NAME} - {datetime.now().strftime('%Y%m%d')}.xlsx"
output_path = os.path.join(OUTPUT_DIR, output_filename)

wb.save(output_path)

# ─────────────────────────────────────────────
# Generate Markdown Summary (for wiki)
# ─────────────────────────────────────────────

markdown_filename = f"Fee Analysis Summary - {SERIES_NAME} - {datetime.now().strftime('%Y%m%d')}.md"
wiki_dir = Path(OUTPUT_DIR).parent / 'wiki'
wiki_dir.mkdir(parents=True, exist_ok=True)
markdown_path = wiki_dir / markdown_filename

# Build markdown table of portfolio fee summaries
md_lines = [
    f"# Fee Analysis Summary — {SERIES_NAME}",
    f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    "",
    "## Fee Summary by Portfolio",
    "",
    "| Portfolio | IM Fee % p.a. | RE Fee % p.a. | Holdings | Status |",
    "|---|---|---|---|---|",
]

for p in PORTFOLIOS:
    im_fee = p.get('im_fee_bps', 0)
    re_fee = p.get('re_fee_bps', 0)
    n_holdings = len(p.get('holdings', []))
    md_lines.append(f"| {p['portfolio_name']} | {im_fee:.2f}% | {re_fee:.2f}% | {n_holdings} | Calculated |")

md_lines.extend([
    "",
    f"**Workbook:** `{output_filename}`",
    "",
    "See the Excel workbook for detailed fee component breakdown and formulas.",
])

with open(markdown_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(md_lines))

# Console output
print(f"✓ Fee analysis workbook generated successfully")
print(f"  Excel: {output_path}")
print(f"  Markdown: {markdown_path}")