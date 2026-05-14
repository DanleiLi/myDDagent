"""
Spreadsheet to JSON Converter
Handles .xlsx, .csv, .tsv files
"""

import json
import csv
from pathlib import Path
from typing import List, Dict, Any
from io import StringIO
from data_cleaner import DataCleaner

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class SpreadsheetConverter:
    """Converts spreadsheets to JSON format."""

    def __init__(self):
        self.output_dir = Path('/mnt/user-data/outputs')
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.cleaner = DataCleaner()
    
    def convert(self, input_path: Path) -> Path:
        """Convert spreadsheet to JSON."""
        file_ext = input_path.suffix.lower()

        if file_ext in ['.xlsx', '.xlsm']:
            output_paths = self._convert_excel(input_path)
        elif file_ext in ['.csv', '.tsv']:
            records = self._convert_csv(input_path, file_ext)
            result = self.cleaner.clean(records)
            output_path = self.output_dir / f"{input_path.stem}.json"
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, indent=2, ensure_ascii=False)
            output_paths = [output_path]
        else:
            raise ValueError(f"Unsupported spreadsheet format: {file_ext}")

        # Return primary output (first file) for compatibility
        return output_paths[0] if isinstance(output_paths, list) else output_paths
    
    def _convert_excel(self, input_path: Path) -> List[Path]:
        """Convert Excel file to JSON. Creates separate files for each sheet."""
        if pd is not None:
            return self._convert_excel_pandas(input_path)
        elif load_workbook is not None:
            return self._convert_excel_openpyxl(input_path)
        else:
            raise ImportError("pandas or openpyxl required. Install with: pip install pandas openpyxl")
    
    def _convert_excel_pandas(self, input_path: Path) -> List[Path]:
        """Convert Excel using pandas. Creates separate JSON file for each sheet."""
        excel_file = pd.ExcelFile(input_path)
        output_paths = []

        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(input_path, sheet_name=sheet_name)
            records = self._dataframe_to_json(df)
            result = self.cleaner.clean(records)

            # Create file for each sheet
            safe_sheet_name = "".join(c if c.isalnum() or c in "_ -" else "_" for c in sheet_name)
            output_path = self.output_dir / f"{input_path.stem}_{safe_sheet_name}.json"

            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, indent=2, ensure_ascii=False)

            output_paths.append(output_path)

        return output_paths
    
    def _convert_excel_openpyxl(self, input_path: Path) -> List[Path]:
        """Convert Excel using openpyxl (fallback). Creates separate JSON file for each sheet."""
        wb = load_workbook(input_path, data_only=True)
        output_paths = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            records = self._worksheet_to_json(ws)
            result = self.cleaner.clean(records)

            # Create file for each sheet
            safe_sheet_name = "".join(c if c.isalnum() or c in "_ -" else "_" for c in sheet_name)
            output_path = self.output_dir / f"{input_path.stem}_{safe_sheet_name}.json"

            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, indent=2, ensure_ascii=False)

            output_paths.append(output_path)

        return output_paths
    
    def _dataframe_to_json(self, df) -> List[Dict[str, Any]]:
        """Convert pandas DataFrame to JSON."""
        # Handle NaN/None values
        df = df.where(pd.notna(df), None)
        
        # Convert to list of dictionaries
        records = df.to_dict(orient='records')
        
        return records
    
    def _worksheet_to_json(self, ws) -> List[Dict[str, Any]]:
        """Convert openpyxl worksheet to JSON."""
        rows = list(ws.iter_rows(values_only=True))
        
        if not rows:
            return []
        
        # Use first row as headers
        headers = rows[0]
        
        result = []
        for row in rows[1:]:
            record = {}
            for header, value in zip(headers, row):
                if header is not None:
                    # Type conversion
                    record[str(header)] = self._convert_value(value)
            result.append(record)
        
        return result
    
    def _convert_csv(self, input_path: Path, file_ext: str) -> List[Dict[str, Any]]:
        """Convert CSV/TSV to JSON."""
        delimiter = '\t' if file_ext == '.tsv' else ','
        
        # Try with pandas first for better type inference
        if pd is not None:
            try:
                df = pd.read_csv(input_path, delimiter=delimiter)
                return self._dataframe_to_json(df)
            except Exception:
                pass
        
        # Fallback to manual CSV parsing
        return self._parse_csv_manual(input_path, delimiter)
    
    def _parse_csv_manual(self, input_path: Path, delimiter: str) -> List[Dict[str, Any]]:
        """Parse CSV manually with type inference."""
        result = []
        
        with open(input_path, 'r', encoding='utf-8', errors='replace') as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            
            for row in reader:
                # Convert values
                converted_row = {}
                for key, value in row.items():
                    converted_row[key] = self._convert_value(value)
                result.append(converted_row)
        
        return result
    
    def _convert_value(self, value) -> Any:
        """Convert cell value to appropriate JSON type."""
        if value is None or value == '':
            return None
        
        if isinstance(value, (int, float, bool)):
            return value
        
        if isinstance(value, str):
            # Try to convert to number
            if value.lower() in ['true', 'yes']:
                return True
            elif value.lower() in ['false', 'no']:
                return False
            
            try:
                if '.' in value:
                    return float(value)
                else:
                    return int(value)
            except ValueError:
                pass
        
        # Return as string
        return str(value) if value is not None else None
